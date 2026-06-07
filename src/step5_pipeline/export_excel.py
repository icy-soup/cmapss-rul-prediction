"""把所有实验结果汇总到一张 Excel 里，方便汇报和归档"""
import json, numpy as np
from pathlib import Path
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

RESULTS = Path("results")
wb = openpyxl.Workbook()

# 表格样式
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT_W = Font(bold=True, size=11, color="FFFFFF")
BEST_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# 从模型定义里查到的参数量，懒得每次重算
PARAMS = {
    "PatchiTransformerRUL": {"FD001": 1748561, "FD002": 1945169, "FD003": 1748561, "FD004": 1945169},
    "ConvPatchiTransformerRUL": {"FD001": 1748566, "FD002": 1945174, "FD003": 1748566, "FD004": 1945174},
    "MSPatchiTransformerRUL": {"FD001": 3135840, "FD002": 3791200, "FD003": 3135840, "FD004": 3791200},
}

def _find(pattern: str) -> Path:
    """在 results 目录树里找文件，主实验和消融子目录都会翻一遍"""
    p = RESULTS / pattern
    if p.exists(): return p
    for sub in ["main_experiments", "ablation"]:
        p = RESULTS / sub / pattern
        if p.exists(): return p
    return RESULTS / pattern

def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT_W
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

def auto_width(ws, ncols, nrows):
    """自动调整列宽，别太宽也别太窄"""
    for col in range(1, ncols + 1):
        max_len = 0
        for row in range(1, nrows + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 25)

# ===== 第一个工作表：三个模型四个子集的成绩总表 =====
ws = wb.active
ws.title = "Summary"
headers = ["Model", "Subset", "Window", "RMSE", "Score", "R2", "Params", "Status"]
ws.append(headers)
style_header(ws, len(headers))

models = ["PatchiTransformerRUL", "ConvPatchiTransformerRUL", "MSPatchiTransformerRUL"]
subsets = ["FD001", "FD002", "FD003", "FD004"]
row = 2
for model in models:
    for subset in subsets:
        f = _find(f"{model}_{subset}_metrics.json")
        if not f.exists():
            continue
        m = json.load(open(f))
        # 从配置文件里翻一下窗口大小
        cfg_f = _find(f"{model}_{subset}_config.json")
        cfg = json.load(open(cfg_f)) if cfg_f.exists() else {}
        wsize = cfg.get("window_size", "?")
        params = f"{PARAMS[model][subset]:,}"
        status = "OK"

        ws.append([model, subset, wsize,
                    round(m["RMSE"], 2), round(m["Score"], 2), round(m.get("R2", 0), 4),
                    params, status])
        # Person 2 是核心模型，给行标个绿色
        if "MSPatch" in model:
            for c in range(1, 9):
                ws.cell(row=row, column=c).fill = BEST_FILL
        row += 1

auto_width(ws, len(headers), row - 1)

# ===== 第二个工作表：FD001 上的消融实验 =====
ws2 = wb.create_sheet("Ablation FD001")
headers2 = ["Experiment", "Branches", "RMSE", "R2", "Score", "Params"]
ws2.append(headers2)
style_header(ws2, len(headers2))

ablation_csv = _find("ablation_summary.csv")
row2 = 2
if ablation_csv.exists():
    with open(ablation_csv) as f:
        reader = csv.DictReader(f)
        for r in sorted(reader, key=lambda x: float(x["RMSE"])):
            ws2.append([r["experiment"], r["branches"],
                        round(float(r["RMSE"]), 2), round(float(r["R2"]), 4),
                        round(float(r["Score"]), 1), r["params"]])
            row2 += 1
auto_width(ws2, len(headers2), row2 - 1)

# ===== 第三个工作表：FD004 上的消融验证 =====
ws3 = wb.create_sheet("Ablation FD004")
headers3 = ["Config", "RMSE", "R2", "Score", "Params"]
ws3.append(headers3)
style_header(ws3, len(headers3))

# 逐条读 FD004 消融的实验结果文件
fd004_configs = [
    ("ablation_FD004_single-small_metrics.json", "P8S4"),
    ("ablation_FD004_single-medium_metrics.json", "P16S8"),
    ("ablation_FD004_small+medium_metrics.json", "P8S4+P16S8"),
]
fd004_pm = {"P8S4": "2.53M", "P16S8": "2.14M", "P8S4+P16S8": "3.13M"}
row3 = 2
for fname, label in fd004_configs:
    p = _find(fname)
    if p.exists():
        m = json.load(open(p))
        ws3.append([label, round(m["RMSE"], 2), round(m["R2"], 4),
                    round(m["Score"], 1), fd004_pm[label]])
        row3 += 1

# 再从主实验里拿三分支和 Base 的 FD004 结果做个对比参考
for model, subset, label, pm in [
    ("MSPatchiTransformerRUL", "FD004", "3-branch", "3.79M"),
    ("PatchiTransformerRUL", "FD004", "Base", "1.95M"),
]:
    p = _find(f"{model}_{subset}_metrics.json")
    if p.exists():
        m = json.load(open(p))
        ws3.append([label, round(m["RMSE"], 2), round(m["R2"], 4),
                    round(m["Score"], 1), pm])
        row3 += 1

auto_width(ws3, len(headers3), row3 - 1)

# ===== 第四个工作表：训练历史摘要 =====
ws4 = wb.create_sheet("Training")
headers4 = ["Model", "Subset", "Epochs", "Final Loss", "Best Val RMSE"]
ws4.append(headers4)
style_header(ws4, len(headers4))
row4 = 2
for model in models:
    for subset in subsets:
        f = _find(f"{model}_{subset}_history.npz")
        if not f.exists():
            continue
        d = np.load(f)
        epochs = len(d["train_loss"])
        final_loss = float(d["train_loss"][-1])
        best_val_rmse = float(min(d["val_rmse"])) if "val_rmse" in d else None
        ws4.append([model, subset, epochs, round(final_loss, 2),
                    round(best_val_rmse, 2) if best_val_rmse else ""])
        row4 += 1
auto_width(ws4, len(headers4), row4 - 1)

xlsx_path = RESULTS / "CMAPSS_Results.xlsx"
wb.save(xlsx_path)
print(f"Excel 已保存到 {xlsx_path}")
print(f"工作表: {wb.sheetnames}")
